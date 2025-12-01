#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Расчет вентиляции в горячем цехе и обеденном зале кафе
Пример 5.1 из Сборника расчетов АВОК-2020
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_ventilation_calculation():
    wb = Workbook()

    # Удаляем стандартный лист
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # Создаем листы
    ws_main = wb.create_sheet("Расчет вентиляции")
    ws_tables = wb.create_sheet("Справочные таблицы")

    # Стили
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    subheader_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
    subheader_font = Font(bold=True, size=10)
    result_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    result_font = Font(bold=True, size=11, color="C65911")

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ===================== ОСНОВНОЙ ЛИСТ РАСЧЕТА =====================
    ws = ws_main
    row = 1

    # Заголовок
    ws.merge_cells(f'A{row}:E{row}')
    cell = ws[f'A{row}']
    cell.value = "РАСЧЕТ ВЕНТИЛЯЦИИ В ГОРЯЧЕМ ЦЕХЕ И ОБЕДЕННОМ ЗАЛЕ КАФЕ"
    cell.font = Font(bold=True, size=14, color="FFFFFF")
    cell.fill = PatternFill(start_color="203764", end_color="203764", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    row += 1

    ws.merge_cells(f'A{row}:E{row}')
    cell = ws[f'A{row}']
    cell.value = "Пример 5.1 из Сборника расчетов АВОК-2020"
    cell.font = Font(italic=True, size=10)
    cell.alignment = Alignment(horizontal="center")
    row += 2

    # ИСХОДНЫЕ ДАННЫЕ
    ws.merge_cells(f'A{row}:E{row}')
    cell = ws[f'A{row}']
    cell.value = "ИСХОДНЫЕ ДАННЫЕ"
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    row += 1

    # Общие данные
    ws[f'A{row}'] = "Параметр"
    ws[f'B{row}'] = "Обозначение"
    ws[f'C{row}'] = "Значение"
    ws[f'D{row}'] = "Единица"
    ws[f'E{row}'] = "Примечание"
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws[f'{col}{row}'].font = subheader_font
        ws[f'{col}{row}'].fill = subheader_fill
        ws[f'{col}{row}'].border = thin_border
        ws[f'{col}{row}'].alignment = Alignment(horizontal="center", vertical="center")
    row += 1

    # Данные
    initial_data = [
        ("Количество посадочных мест", "n₁", 20, "мест", "Обеденный зал"),
        ("Численность персонала в горячем цехе", "n₂", 3, "чел", ""),
        ("Объем горячего цеха", "Vг", 90, "м³", "Высота 3,1 м"),
        ("Объем обеденного зала", "Vт", 331.2, "м³", "Высота 3 м"),
        ("Коэффициент одновременности", "Kо", 0.7, "-", ""),
        ("Коэффициент эффективности отсоса", "Kэфобщ", 0.7, "-", "Вытяжные зонты"),
        ("Мощность освещения горячего цеха", "Q₃", 0.71, "кВт", ""),
        ("Теплопоступления от солн. радиации (гор. цех)", "Q₄", 0, "кВт", "Окон нет"),
        ("Мощность освещения обеденного зала", "Q₃т", 1.66, "кВт", ""),
        ("Теплопоступления от солн. радиации (обед. зал)", "Q₄т", 4.27, "кВт", ""),
        ("Температура воздуха, удаляемого отсосами", "tотс", 42, "°C", ""),
        ("Температура воздуха в верхней зоне", "tв", 30, "°C", ""),
        ("Температура воздуха в обеденном зале", "tрт", 27, "°C", ""),
        ("Кратность воздухообмена", "n", 2, "ч⁻¹", ""),
    ]

    start_data_row = row
    for i, (param, symbol, value, unit, note) in enumerate(initial_data):
        ws[f'A{row}'] = param
        ws[f'B{row}'] = symbol
        ws[f'C{row}'] = value
        ws[f'D{row}'] = unit
        ws[f'E{row}'] = note
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws[f'{col}{row}'].border = thin_border
            ws[f'{col}{row}'].alignment = Alignment(horizontal="left" if col in ['A', 'E'] else "center", vertical="center")
        row += 1

    row += 1

    # ТАБЛИЦА 5.1 - Оборудование горячего цеха
    ws.merge_cells(f'A{row}:H{row}')
    cell = ws[f'A{row}']
    cell.value = "ТАБЛИЦА 5.1 - Перечень теплового оборудования в горячем цехе"
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    row += 1

    headers_51 = ["Оборудование", "Кол-во, ед.", "Qу, кВт", "Kз", "Вытяжка, м³/ч", "Приток, м³/ч", "Qу·Kз", "Примечание"]
    for i, header in enumerate(headers_51):
        col_letter = get_column_letter(i + 1)
        ws[f'{col_letter}{row}'] = header
        ws[f'{col_letter}{row}'].font = subheader_font
        ws[f'{col_letter}{row}'].fill = subheader_fill
        ws[f'{col_letter}{row}'].border = thin_border
        ws[f'{col_letter}{row}'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    row += 1

    table51_start = row
    equipment_data = [
        ("Пароконвекционная печь", 2, 23, 0.65, 1700, 400, "С отсосом"),
        ("Пароконвекционная печь", 2, 17.5, 0.65, 1500, 400, "С отсосом"),
        ("Холодильный прилавок", 2, 0.34, 0.3, "", "", "Без отсоса"),
        ("Гриль", 1, 3.6, 0.3, "", "", "Без отсоса"),
        ("Станция напитков", 1, 1, 0.3, "", "", "Без отсоса"),
        ("Охладитель молока", 1, 0.5, 0.3, "", "", "Без отсоса"),
        ("Льдогенератор", 1, 1, 0.3, "", "", "Без отсоса"),
        ("Слайсер", 1, 0.25, 0.3, "", "", "Без отсоса"),
    ]

    for equip, count, power, kz, exhaust, supply, note in equipment_data:
        ws[f'A{row}'] = equip
        ws[f'B{row}'] = count
        ws[f'C{row}'] = power
        ws[f'D{row}'] = kz
        ws[f'E{row}'] = exhaust
        ws[f'F{row}'] = supply
        ws[f'G{row}'].value = f"=B{row}*C{row}*D{row}"  # Формула для Qу·Kз
        ws[f'H{row}'] = note

        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws[f'{col}{row}'].border = thin_border
            ws[f'{col}{row}'].alignment = Alignment(horizontal="left" if col in ['A', 'H'] else "center", vertical="center")
        row += 1

    table51_end = row - 1
    row += 1

    # РАСЧЕТЫ
    ws.merge_cells(f'A{row}:E{row}')
    cell = ws[f'A{row}']
    cell.value = "РАСЧЕТЫ"
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    row += 1

    # Формула (1) - Тепловыделения от оборудования
    ws[f'A{row}'] = "1. ТЕПЛОВЫДЕЛЕНИЯ ОТ КУХОННОГО ОБОРУДОВАНИЯ (Формула 1)"
    ws[f'A{row}'].font = Font(bold=True, size=11)
    row += 1

    ws[f'A{row}'] = "Описание"
    ws[f'C{row}'] = "Формула/Значение"
    ws[f'D{row}'] = "Результат"
    ws[f'E{row}'] = "Единица"
    for col in ['A', 'C', 'D', 'E']:
        ws[f'{col}{row}'].font = subheader_font
        ws[f'{col}{row}'].fill = subheader_fill
        ws[f'{col}{row}'].border = thin_border
    row += 1

    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'] = "Оборудование с местными отсосами (∑Qу·Kз)"
    ws[f'C{row}'] = f"=SUMIF(E{table51_start}:E{table51_end},\">0\",G{table51_start}:G{table51_end})"
    ws[f'D{row}'].value = f"=C{row}"
    ws[f'E{row}'] = "кВт"
    for col in ['A', 'C', 'D', 'E']:
        ws[f'{col}{row}'].border = thin_border
    sum_with_exhaust_row = row
    row += 1

    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'] = "С учетом эффективности отсоса (1 - Kэфобщ)"
    ws[f'C{row}'] = f"=D{sum_with_exhaust_row}*(1-C{start_data_row+5})"
    ws[f'D{row}'].value = f"=C{row}"
    ws[f'E{row}'] = "кВт"
    for col in ['A', 'C', 'D', 'E']:
        ws[f'{col}{row}'].border = thin_border
    with_efficiency_row = row
    row += 1

    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'] = "Оборудование без местных отсосов (∑Qу_р·Kз)"
    ws[f'C{row}'] = f"=SUMIF(E{table51_start}:E{table51_end},\"\",G{table51_start}:G{table51_end})"
    ws[f'D{row}'].value = f"=C{row}"
    ws[f'E{row}'] = "кВт"
    for col in ['A', 'C', 'D', 'E']:
        ws[f'{col}{row}'].border = thin_border
    without_exhaust_row = row
    row += 1

    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'] = "Q₁ = Kо·[(∑Qу·Kз·(1-Kэфобщ)) + (∑Qу_р·Kз)]"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'C{row}'] = f"=C{start_data_row+4}*(D{with_efficiency_row}+D{without_exhaust_row})"
    ws[f'D{row}'].value = f"=C{row}"
    ws[f'D{row}'].fill = result_fill
    ws[f'D{row}'].font = result_font
    ws[f'E{row}'] = "кВт"
    for col in ['A', 'C', 'D', 'E']:
        ws[f'{col}{row}'].border = thin_border
    Q1_row = row
    row += 2

    # Формула (2) - Тепловыделения от людей
    ws[f'A{row}'] = "2. ТЕПЛОВЫДЕЛЕНИЯ ОТ ЛЮДЕЙ (Формула 2)"
    ws[f'A{row}'].font = Font(bold=True, size=11)
    row += 1

    ws[f'A{row}'] = "Описание"
    ws[f'C{row}'] = "Формула/Значение"
    ws[f'D{row}'] = "Результат"
    ws[f'E{row}'] = "Единица"
    for col in ['A', 'C', 'D', 'E']:
        ws[f'{col}{row}'].font = subheader_font
        ws[f'{col}{row}'].fill = subheader_fill
        ws[f'{col}{row}'].border = thin_border
    row += 1

    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'] = "Численность персонала (n₂)"
    ws[f'C{row}'] = f"=C{start_data_row+1}"
    ws[f'D{row}'].value = f"=C{row}"
    ws[f'E{row}'] = "чел"
    for col in ['A', 'C', 'D', 'E']:
        ws[f'{col}{row}'].border = thin_border
    row += 1

    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'] = "Тепловыделения от 1 работающего (q₂)"
    ws[f'C{row}'] = 0.2
    ws[f'D{row}'].value = f"=C{row}"
    ws[f'E{row}'] = "кВт"
    for col in ['A', 'C', 'D', 'E']:
        ws[f'{col}{row}'].border = thin_border
    q2_row = row
    row += 1

    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'] = "Q₂ = n₂ · q₂"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'C{row}'] = f"=C{start_data_row+1}*C{q2_row}"
    ws[f'D{row}'].value = f"=C{row}"
    ws[f'D{row}'].fill = result_fill
    ws[f'D{row}'].font = result_font
    ws[f'E{row}'] = "кВт"
    for col in ['A', 'C', 'D', 'E']:
        ws[f'{col}{row}'].border = thin_border
    Q2_row = row
    row += 2

    # Формула (3) - Общие тепловыделения
    ws[f'A{row}'] = "3. ОБЩИЕ ТЕПЛОВЫДЕЛЕНИЯ В ГОРЯЧЕМ ЦЕХЕ (Формула 3)"
    ws[f'A{row}'].font = Font(bold=True, size=11)
    row += 1

    ws[f'A{row}'] = "Описание"
    ws[f'C{row}'] = "Формула/Значение"
    ws[f'D{row}'] = "Результат"
    ws[f'E{row}'] = "Единица"
    for col in ['A', 'C', 'D', 'E']:
        ws[f'{col}{row}'].font = subheader_font
        ws[f'{col}{row}'].fill = subheader_fill
        ws[f'{col}{row}'].border = thin_border
    row += 1

    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'] = "Тепловыделения от оборудования (Q₁)"
    ws[f'C{row}'] = f"=D{Q1_row}"
    ws[f'D{row}'].value = f"=C{row}"
    ws[f'E{row}'] = "кВт"
    for col in ['A', 'C', 'D', 'E']:
        ws[f'{col}{row}'].border = thin_border
    row += 1

    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'] = "Тепловыделения от людей (Q₂)"
    ws[f'C{row}'] = f"=D{Q2_row}"
    ws[f'D{row}'].value = f"=C{row}"
    ws[f'E{row}'] = "кВт"
    for col in ['A', 'C', 'D', 'E']:
        ws[f'{col}{row}'].border = thin_border
    row += 1

    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'] = "Тепловыделения от освещения (Q₃)"
    ws[f'C{row}'] = f"=C{start_data_row+6}"
    ws[f'D{row}'].value = f"=C{row}"
    ws[f'E{row}'] = "кВт"
    for col in ['A', 'C', 'D', 'E']:
        ws[f'{col}{row}'].border = thin_border
    row += 1

    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'] = "Qг = Q₁ + Q₂ + Q₃"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'C{row}'] = f"=D{Q1_row}+D{Q2_row}+C{start_data_row+6}"
    ws[f'D{row}'].value = f"=C{row}"
    ws[f'D{row}'].fill = result_fill
    ws[f'D{row}'].font = result_font
    ws[f'E{row}'] = "кВт"
    for col in ['A', 'C', 'D', 'E']:
        ws[f'{col}{row}'].border = thin_border
    Qg_row = row
    row += 2

    # Формула (4) - Воздух из верхней зоны
    ws[f'A{row}'] = "4. ВОЗДУХ, УДАЛЯЕМЫЙ ИЗ ВЕРХНЕЙ ЗОНЫ (Формула 4)"
    ws[f'A{row}'].font = Font(bold=True, size=11)
    row += 1

    ws[f'A{row}'] = "Описание"
    ws[f'C{row}'] = "Формула/Значение"
    ws[f'D{row}'] = "Результат"
    ws[f'E{row}'] = "Единица"
    for col in ['A', 'C', 'D', 'E']:
        ws[f'{col}{row}'].font = subheader_font
        ws[f'{col}{row}'].fill = subheader_fill
        ws[f'{col}{row}'].border = thin_border
    row += 1

    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'] = "Lв = n · Vг"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'C{row}'] = f"=C{start_data_row+13}*C{start_data_row+2}"
    ws[f'D{row}'].value = f"=C{row}"
    ws[f'D{row}'].fill = result_fill
    ws[f'D{row}'].font = result_font
    ws[f'E{row}'] = "м³/ч"
    for col in ['A', 'C', 'D', 'E']:
        ws[f'{col}{row}'].border = thin_border
    Lv_row = row
    row += 2

    # Плотности воздуха
    ws[f'A{row}'] = "5. ПЛОТНОСТИ ВОЗДУХА"
    ws[f'A{row}'].font = Font(bold=True, size=11)
    row += 1

    ws[f'A{row}'] = "Описание"
    ws[f'C{row}'] = "Формула/Значение"
    ws[f'D{row}'] = "Результат"
    ws[f'E{row}'] = "Единица"
    for col in ['A', 'C', 'D', 'E']:
        ws[f'{col}{row}'].font = subheader_font
        ws[f'{col}{row}'].fill = subheader_fill
        ws[f'{col}{row}'].border = thin_border
    row += 1

    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'] = "ρᵢ - для местных отсосов (как в PDF)"
    ws[f'C{row}'] = 1.13  # В PDF используется округленное значение 1,13
    ws[f'D{row}'].value = f"=C{row}"
    ws[f'D{row}'].fill = result_fill
    ws[f'E{row}'] = "кг/м³"
    ws[f'A{row}'].alignment = Alignment(horizontal="left", vertical="center")
    for col in ['A', 'C', 'D', 'E']:
        ws[f'{col}{row}'].border = thin_border
    rho_i_row = row
    row += 1

    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'] = "ρв = 353/(273 + tв) - для верхней зоны"
    ws[f'C{row}'] = f"=353/(273+C{start_data_row+11})"
    ws[f'D{row}'].value = f"=C{row}"
    ws[f'D{row}'].fill = result_fill
    ws[f'E{row}'] = "кг/м³"
    for col in ['A', 'C', 'D', 'E']:
        ws[f'{col}{row}'].border = thin_border
    rho_v_row = row
    row += 2

    # Формула (5) - Массовый расход удаляемого воздуха
    ws[f'A{row}'] = "6. МАССОВЫЙ РАСХОД УДАЛЯЕМОГО ВОЗДУХА (Формула 5)"
    ws[f'A{row}'].font = Font(bold=True, size=11)
    row += 1

    ws[f'A{row}'] = "Описание"
    ws[f'C{row}'] = "Формула/Значение"
    ws[f'D{row}'] = "Результат"
    ws[f'E{row}'] = "Единица"
    for col in ['A', 'C', 'D', 'E']:
        ws[f'{col}{row}'].font = subheader_font
        ws[f'{col}{row}'].fill = subheader_fill
        ws[f'{col}{row}'].border = thin_border
    row += 1

    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'] = "∑Lоᵢ - суммарный расход через местные отсосы"
    ws[f'C{row}'] = f"=SUMIF(E{table51_start}:E{table51_end},\">0\",E{table51_start}:E{table51_end})*2"
    ws[f'D{row}'].value = f"=C{row}"
    ws[f'E{row}'] = "м³/ч"
    for col in ['A', 'C', 'D', 'E']:
        ws[f'{col}{row}'].border = thin_border
    sum_Lo_row = row
    row += 1

    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'] = "Gуг = (∑Lоᵢ · ρᵢ) + Lв · ρв"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'C{row}'] = f"=D{sum_Lo_row}*D{rho_i_row}+D{Lv_row}*D{rho_v_row}"
    ws[f'D{row}'].value = f"=C{row}"
    ws[f'D{row}'].fill = result_fill
    ws[f'D{row}'].font = result_font
    ws[f'E{row}'] = "кг/ч"
    for col in ['A', 'C', 'D', 'E']:
        ws[f'{col}{row}'].border = thin_border
    Guy_row = row
    row += 2

    # Расходы воздуха
    ws[f'A{row}'] = "7. РАСПРЕДЕЛЕНИЕ ВОЗДУШНЫХ ПОТОКОВ"
    ws[f'A{row}'].font = Font(bold=True, size=11)
    row += 1

    ws[f'A{row}'] = "Описание"
    ws[f'C{row}'] = "Формула/Значение"
    ws[f'D{row}'] = "Результат"
    ws[f'E{row}'] = "Единица"
    for col in ['A', 'C', 'D', 'E']:
        ws[f'{col}{row}'].font = subheader_font
        ws[f'{col}{row}'].fill = subheader_fill
        ws[f'{col}{row}'].border = thin_border
    row += 1

    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'] = "Gс = 0,2 · Gуг (через раздаточный проем)"
    ws[f'C{row}'] = f"=0.2*D{Guy_row}"
    ws[f'D{row}'].value = f"=C{row}"
    ws[f'D{row}'].fill = result_fill
    ws[f'D{row}'].font = result_font
    ws[f'E{row}'] = "кг/ч"
    for col in ['A', 'C', 'D', 'E']:
        ws[f'{col}{row}'].border = thin_border
    Gc_row = row
    row += 1

    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'] = "Gпог = 0,8 · Gуг (приточный воздух в гор. цех)"
    ws[f'C{row}'] = f"=0.8*D{Guy_row}"
    ws[f'D{row}'].value = f"=C{row}"
    ws[f'D{row}'].fill = result_fill
    ws[f'D{row}'].font = result_font
    ws[f'E{row}'] = "кг/ч"
    for col in ['A', 'C', 'D', 'E']:
        ws[f'{col}{row}'].border = thin_border
    Gpog_row = row
    row += 2

    # Плотность воздуха в обеденном зале
    ws[f'A{row}'] = "8. ОБЪЕМНЫЙ РАСХОД ЧЕРЕЗ РАЗДАТОЧНЫЙ ПРОЕМ"
    ws[f'A{row}'].font = Font(bold=True, size=11)
    row += 1

    ws[f'A{row}'] = "Описание"
    ws[f'C{row}'] = "Формула/Значение"
    ws[f'D{row}'] = "Результат"
    ws[f'E{row}'] = "Единица"
    for col in ['A', 'C', 'D', 'E']:
        ws[f'{col}{row}'].font = subheader_font
        ws[f'{col}{row}'].fill = subheader_fill
        ws[f'{col}{row}'].border = thin_border
    row += 1

    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'] = "ρрт = 353/(273 + tрт) - плотность в обед. зале"
    ws[f'C{row}'] = f"=353/(273+C{start_data_row+12})"
    ws[f'D{row}'].value = f"=C{row}"
    ws[f'D{row}'].fill = result_fill
    ws[f'E{row}'] = "кг/м³"
    for col in ['A', 'C', 'D', 'E']:
        ws[f'{col}{row}'].border = thin_border
    rho_rt_row = row
    row += 1

    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'] = "Lс = Gс / ρрт"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'C{row}'] = f"=D{Gc_row}/D{rho_rt_row}"
    ws[f'D{row}'].value = f"=C{row}"
    ws[f'D{row}'].fill = result_fill
    ws[f'D{row}'].font = result_font
    ws[f'E{row}'] = "м³/ч"
    for col in ['A', 'C', 'D', 'E']:
        ws[f'{col}{row}'].border = thin_border
    Lc_row = row
    row += 2

    # ИТОГОВЫЕ РЕЗУЛЬТАТЫ
    ws.merge_cells(f'A{row}:E{row}')
    cell = ws[f'A{row}']
    cell.value = "ИТОГОВЫЕ РЕЗУЛЬТАТЫ"
    cell.font = Font(bold=True, size=12, color="FFFFFF")
    cell.fill = PatternFill(start_color="C65911", end_color="C65911", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    row += 1

    results = [
        ("Тепловыделения от оборудования Q₁", f"=D{Q1_row}", "кВт", "12,53 кВт"),
        ("Тепловыделения от людей Q₂", f"=D{Q2_row}", "кВт", "0,6 кВт"),
        ("Общие тепловыделения Qг", f"=D{Qg_row}", "кВт", "13,84 кВт"),
        ("Воздух из верхней зоны Lв", f"=D{Lv_row}", "м³/ч", "180 м³/ч"),
        ("Плотность воздуха в отсосах ρᵢ", f"=D{rho_i_row}", "кг/м³", "1,13 кг/м³"),
        ("Плотность воздуха верхней зоны ρв", f"=D{rho_v_row}", "кг/м³", "1,17 кг/м³"),
        ("Массовый расход удаляемого воздуха Gуг", f"=D{Guy_row}", "кг/ч", "7440 кг/ч"),
        ("Расход через раздаточный проем Gс", f"=D{Gc_row}", "кг/ч", "1490 кг/ч"),
        ("Приточный воздух в горячий цех Gпог", f"=D{Gpog_row}", "кг/ч", "5950 кг/ч"),
        ("Плотность в обеденном зале ρрт", f"=D{rho_rt_row}", "кг/м³", "1,18 кг/м³"),
        ("Объемный расход через проем Lс", f"=D{Lc_row}", "м³/ч", "1260 м³/ч"),
    ]

    ws[f'A{row}'] = "Параметр"
    ws[f'C{row}'] = "Значение"
    ws[f'D{row}'] = "Единица"
    ws[f'E{row}'] = "Контроль (из PDF)"
    for col in ['A', 'C', 'D', 'E']:
        ws[f'{col}{row}'].font = Font(bold=True)
        ws[f'{col}{row}'].fill = subheader_fill
        ws[f'{col}{row}'].border = thin_border
        ws[f'{col}{row}'].alignment = Alignment(horizontal="center", vertical="center")
    row += 1

    for param, formula, unit, control in results:
        ws[f'A{row}'] = param
        ws.merge_cells(f'A{row}:B{row}')
        ws[f'C{row}'].value = formula
        ws[f'C{row}'].number_format = '0.00'
        ws[f'D{row}'] = unit
        ws[f'E{row}'] = control

        for col in ['A', 'C', 'D', 'E']:
            ws[f'{col}{row}'].border = thin_border
            ws[f'{col}{row}'].alignment = Alignment(horizontal="left" if col == 'A' else "center", vertical="center")

        ws[f'C{row}'].fill = result_fill
        ws[f'C{row}'].font = Font(bold=True, color="C65911")
        row += 1

    # Настройка ширины столбцов
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15

    # ===================== ЛИСТ СО СПРАВОЧНЫМИ ТАБЛИЦАМИ =====================
    ws_ref = ws_tables
    row = 1

    # Заголовок
    ws_ref.merge_cells(f'A{row}:E{row}')
    cell = ws_ref[f'A{row}']
    cell.value = "СПРАВОЧНЫЕ ТАБЛИЦЫ"
    cell.font = Font(bold=True, size=14, color="FFFFFF")
    cell.fill = PatternFill(start_color="203764", end_color="203764", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    row += 2

    # Таблица 5.4 - Тепловыделения от людей
    ws_ref.merge_cells(f'A{row}:H{row}')
    cell = ws_ref[f'A{row}']
    cell.value = "ТАБЛИЦА 5.4 - Количество теплоты, выделяемое взрослыми мужчинами"
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    row += 1

    # Заголовки
    ws_ref.merge_cells(f'A{row}:A{row+1}')
    ws_ref[f'A{row}'] = "Показатели"
    ws_ref.merge_cells(f'B{row}:H{row}')
    ws_ref[f'B{row}'] = "Количество теплоты, Вт, при температуре воздуха в помещении, °C"

    for col in ['A', 'B']:
        ws_ref[f'{col}{row}'].font = subheader_font
        ws_ref[f'{col}{row}'].fill = subheader_fill
        ws_ref[f'{col}{row}'].border = thin_border
        ws_ref[f'{col}{row}'].alignment = Alignment(horizontal="center", vertical="center")
    row += 1

    temps = ["10", "15", "20", "25", "30", "35"]
    for i, temp in enumerate(temps):
        col_letter = get_column_letter(i + 2)
        ws_ref[f'{col_letter}{row}'] = temp
        ws_ref[f'{col_letter}{row}'].font = subheader_font
        ws_ref[f'{col_letter}{row}'].fill = subheader_fill
        ws_ref[f'{col_letter}{row}'].border = thin_border
        ws_ref[f'{col_letter}{row}'].alignment = Alignment(horizontal="center", vertical="center")
    row += 1

    # Данные таблицы 5.4
    heat_data = [
        ("В состоянии покоя", "", "", "", "", "", ""),
        ("  - явная", 140, 120, 90, 60, 40, 10),
        ("  - полная", 165, 145, 120, 95, 95, 95),
        ("При легкой работе", "", "", "", "", "", ""),
        ("  - явная", 150, 120, 100, 65, 40, 5),
        ("  - полная", 180, 160, 150, 145, 145, 145),
        ("При работе средней тяжести", "", "", "", "", "", ""),
        ("  - явная", 165, 135, 105, 70, 40, 5),
        ("  - полная", 215, 210, 205, 200, 200, 200),
        ("При тяжелой работе", "", "", "", "", "", ""),
        ("  - явная", 200, 165, 130, 95, 50, 10),
        ("  - полная", 290, 290, 290, 290, 290, 290),
    ]

    for data_row in heat_data:
        ws_ref[f'A{row}'] = data_row[0]
        for i in range(1, 7):
            col_letter = get_column_letter(i + 1)
            ws_ref[f'{col_letter}{row}'] = data_row[i]
            ws_ref[f'{col_letter}{row}'].border = thin_border
            ws_ref[f'{col_letter}{row}'].alignment = Alignment(horizontal="center", vertical="center")

        ws_ref[f'A{row}'].border = thin_border
        ws_ref[f'A{row}'].alignment = Alignment(horizontal="left", vertical="center")

        if not data_row[1]:  # Если это заголовок категории
            ws_ref[f'A{row}'].font = Font(bold=True)
            ws_ref.merge_cells(f'B{row}:H{row}')

        row += 1

    row += 1

    # Примечание
    ws_ref.merge_cells(f'A{row}:H{row}')
    ws_ref[f'A{row}'] = "* Для женщин эти значения необходимо умножать на 0,85, для детей – на 0,75"
    ws_ref[f'A{row}'].font = Font(italic=True, size=9)
    ws_ref[f'A{row}'].alignment = Alignment(horizontal="left", vertical="center")
    row += 2

    # Таблица 5.5 - Поправка на положение оборудования
    ws_ref.merge_cells(f'A{row}:C{row}')
    cell = ws_ref[f'A{row}']
    cell.value = "ТАБЛИЦА 5.5 - Поправка на положение источника теплоты"
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    row += 1

    ws_ref[f'A{row}'] = "Положение кухонного оборудования"
    ws_ref[f'B{row}'] = "Коэффициент r"
    for col in ['A', 'B']:
        ws_ref[f'{col}{row}'].font = subheader_font
        ws_ref[f'{col}{row}'].fill = subheader_fill
        ws_ref[f'{col}{row}'].border = thin_border
        ws_ref[f'{col}{row}'].alignment = Alignment(horizontal="center", vertical="center")
    row += 1

    position_data = [
        ("Свободно стоящее", "1"),
        ("У стены", "0,63 B/A, но не менее 0,63 и не более 1"),
        ("В углу", "0,4"),
    ]

    for position, coef in position_data:
        ws_ref[f'A{row}'] = position
        ws_ref[f'B{row}'] = coef
        for col in ['A', 'B']:
            ws_ref[f'{col}{row}'].border = thin_border
            ws_ref[f'{col}{row}'].alignment = Alignment(horizontal="left" if col == 'A' else "center", vertical="center")
        row += 1

    row += 1

    # Таблица 5.6 - Коэффициент подвижности
    ws_ref.merge_cells(f'A{row}:C{row}')
    cell = ws_ref[f'A{row}']
    cell.value = "ТАБЛИЦА 5.6 - Поправочный коэффициент a"
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    row += 1

    ws_ref[f'A{row}'] = "Способ воздухораспределения"
    ws_ref[f'B{row}'] = "Коэффициент a"
    for col in ['A', 'B']:
        ws_ref[f'{col}{row}'].font = subheader_font
        ws_ref[f'{col}{row}'].fill = subheader_fill
        ws_ref[f'{col}{row}'].border = thin_border
        ws_ref[f'{col}{row}'].alignment = Alignment(horizontal="center", vertical="center")
    row += 1

    ventilation_data = [
        ("Перемешивающая вентиляция, струйная подача:", ""),
        ("  - через приточные решетки на стенах", "1,25"),
        ("  - через плафонные воздухораспределители", "1,2"),
        ("Вытесняющая вентиляция:", ""),
        ("  - на потолке", "1,1"),
        ("  - в рабочей зоне", "1,05"),
    ]

    for method, coef in ventilation_data:
        ws_ref[f'A{row}'] = method
        ws_ref[f'B{row}'] = coef
        for col in ['A', 'B']:
            ws_ref[f'{col}{row}'].border = thin_border
            ws_ref[f'{col}{row}'].alignment = Alignment(horizontal="left" if col == 'A' else "center", vertical="center")

        if not coef:
            ws_ref[f'A{row}'].font = Font(bold=True)

        row += 1

    # Настройка ширины столбцов
    ws_ref.column_dimensions['A'].width = 50
    ws_ref.column_dimensions['B'].width = 15
    ws_ref.column_dimensions['C'].width = 15
    ws_ref.column_dimensions['D'].width = 15
    ws_ref.column_dimensions['E'].width = 15
    ws_ref.column_dimensions['F'].width = 15
    ws_ref.column_dimensions['G'].width = 15
    ws_ref.column_dimensions['H'].width = 15

    # Сохранение файла
    output_file = "/home/user/CLAUDE/Расчет_вентиляции_кафе.xlsx"
    wb.save(output_file)
    print(f"✓ Excel файл успешно создан: {output_file}")
    return output_file

if __name__ == "__main__":
    create_ventilation_calculation()
