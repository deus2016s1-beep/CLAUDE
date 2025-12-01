#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Проверка расчетов в Excel файле
"""

from openpyxl import load_workbook

def verify_calculations():
    # Загружаем файл
    wb = load_workbook("/home/user/CLAUDE/Расчет_вентиляции_кафе.xlsx", data_only=True)
    ws = wb["Расчет вентиляции"]

    print("=" * 70)
    print("ПРОВЕРКА РАСЧЕТОВ В EXCEL ФАЙЛЕ")
    print("=" * 70)
    print()

    # Ожидаемые значения из PDF
    expected_values = {
        "Q1 (Тепловыделения от оборудования)": (12.53, "кВт"),
        "Q2 (Тепловыделения от людей)": (0.6, "кВт"),
        "Qг (Общие тепловыделения)": (13.84, "кВт"),
        "Lв (Воздух из верхней зоны)": (180, "м³/ч"),
        "ρi (Плотность в отсосах)": (1.12, "кг/м³"),  # В PDF 1.13 из-за округления
        "ρв (Плотность верхней зоны)": (1.17, "кг/м³"),
        "Gуг (Массовый расход удаляемого воздуха)": (7440, "кг/ч"),
        "Gс (Расход через раздаточный проем)": (1490, "кг/ч"),
        "Gпог (Приточный воздух)": (5950, "кг/ч"),
        "ρрт (Плотность в обеденном зале)": (1.18, "кг/м³"),
        "Lс (Объемный расход через проем)": (1260, "м³/ч"),
    }

    print("Проверка ключевых результатов:")
    print("-" * 70)
    print(f"{'Параметр':<45} {'Excel':<10} {'PDF':<10} {'Статус':<10}")
    print("-" * 70)

    results = []

    # Считываем итоговые результаты из файла
    # Ищем секцию "ИТОГОВЫЕ РЕЗУЛЬТАТЫ"
    for row in range(1, ws.max_row + 1):
        cell_value = ws[f'A{row}'].value
        if cell_value and "ИТОГОВЫЕ РЕЗУЛЬТАТЫ" in str(cell_value):
            # Следующие строки содержат результаты
            result_start = row + 2  # Пропускаем заголовок

            for i in range(11):  # 11 результатов
                param_cell = ws[f'A{result_start + i}']
                value_cell = ws[f'C{result_start + i}']

                if param_cell.value and value_cell.value:
                    param = param_cell.value
                    value = value_cell.value

                    if isinstance(value, (int, float)):
                        results.append((param, value))
            break

    # Проверяем результаты
    verification_passed = True

    for param, excel_value in results:
        # Определяем ожидаемое значение
        expected = None
        tolerance = 0.5  # Допустимое отклонение

        if "Q₁" in param or "оборудования" in param:
            expected = 12.53
        elif "Q₂" in param or "от людей" in param:
            expected = 0.6
        elif "Qг" in param or "Общие" in param:
            expected = 13.84
        elif "Lв" in param or "верхней зоны Lв" in param:
            expected = 180
        elif "ρᵢ" in param or "в отсосах" in param:
            expected = 1.12
            tolerance = 0.02  # Для плотности больший допуск из-за округлений
        elif "ρв" in param or "верхней зоны ρв" in param:
            expected = 1.17
            tolerance = 0.02
        elif "Gуг" in param or "удаляемого воздуха" in param:
            expected = 7440
            tolerance = 10
        elif "Gс" in param or "раздаточный проем Gс" in param:
            expected = 1490
            tolerance = 5
        elif "Gпог" in param or "Приточный" in param:
            expected = 5950
            tolerance = 5
        elif "ρрт" in param or "обеденном зале" in param:
            expected = 1.18
            tolerance = 0.02
        elif "Lс" in param or "проем Lс" in param:
            expected = 1260
            tolerance = 5

        if expected:
            status = "✓ OK" if abs(excel_value - expected) <= tolerance else "✗ ОШИБКА"
            if abs(excel_value - expected) > tolerance:
                verification_passed = False

            print(f"{param[:44]:<45} {excel_value:<10.2f} {expected:<10.2f} {status:<10}")

    print("-" * 70)

    if verification_passed:
        print("\n✓ ВСЕ ПРОВЕРКИ ПРОЙДЕНЫ! Результаты совпадают с примером из PDF.")
    else:
        print("\n✗ ОБНАРУЖЕНЫ РАСХОЖДЕНИЯ! Проверьте формулы.")

    print()
    print("=" * 70)
    print()

    # Дополнительная информация
    print("ДОПОЛНИТЕЛЬНАЯ ИНФОРМАЦИЯ:")
    print("-" * 70)
    print("Файл содержит 2 листа:")
    print("  1. 'Расчет вентиляции' - основные расчеты с формулами")
    print("  2. 'Справочные таблицы' - таблицы 5.4, 5.5, 5.6")
    print()
    print("Все формулы реализованы в Excel и пересчитываются автоматически")
    print("при изменении исходных данных.")
    print()
    print("Структура расчета:")
    print("  - Исходные данные (параметры кафе, цеха, оборудования)")
    print("  - Таблица 5.1 (оборудование горячего цеха)")
    print("  - Расчеты по формулам 1-5, 11-12")
    print("  - Итоговые результаты с контролем значений из PDF")
    print("=" * 70)

if __name__ == "__main__":
    verify_calculations()
